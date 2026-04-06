from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.http import FileResponse, Http404, HttpResponse
from django.db.models import Q
from .models import *
import os
import pandas as pd
from django.db.models import Avg, Count
import openpyxl

def home(request):
    """Homepage"""
    slides = Slide.objects.filter(is_active=True)[:5]
    
    # Default slides if none exist
    if not slides:
        slides = [
            {
                'title': 'Selamat Datang di Sistem Informasi',
                'description': 'Menghasilkan lulusan yang kompeten, berakhlak mulia, dan berjiwa wirausaha.',
                'image': {'url': 'https://placehold.co/1920x1080/7e22ce/white?text=Sistem+Informasi+UNH'},
                'link': '#',
                'is_active': True
            },
            {
                'title': 'Unggul dalam Teknologi & Bisnis',
                'description': 'Kurikulum berbasis kompetensi yang relevan dengan kebutuhan industri digital masa kini.',
                'image': {'url': 'https://placehold.co/1920x1080/d97706/white?text=Digital+Business'},
                'link': '#',
                'is_active': True
            },
            {
                'title': 'Bergabunglah Bersama Kami',
                'description': 'Jadilah bagian dari komunitas akademik yang dinamis dan inovatif.',
                'image': {'url': 'https://placehold.co/1920x1080/1e40af/white?text=Join+Us'},
                'link': '#',
                'is_active': True
            }
        ]

    sambutan = Sambutan.objects.filter(is_active=True).first()
    news = News.objects.filter(is_published=True)[:6]
    featured_news = News.objects.filter(is_published=True, is_featured=True).first()
    gallery = Gallery.objects.filter(is_active=True)[:8]
    
    context = {
        'slides': slides,
        'sambutan': sambutan,
        'news': news,
        'featured_news': featured_news,
        'gallery': gallery,
    }
    return render(request, 'home.html', context)


def profile(request):
    """Halaman Profile Prodi"""
    visi_misi = VisiMisi.objects.first()
    prospek = ProspekKeunggulan.objects.all()
    
    context = {
        'visi_misi': visi_misi,
        'prospek': prospek,
    }
    return render(request, 'profile.html', context)


def akademik(request):
    """Halaman Akademik"""
    dosen_tetap = Dosen.objects.prefetch_related('mata_kuliah_diampu').filter(kategori='tetap', is_active=True)
    dosen_dtpr = Dosen.objects.prefetch_related('mata_kuliah_diampu').filter(kategori__in=['dtpr'], is_active=True)
    dosen_lb = Dosen.objects.prefetch_related('mata_kuliah_diampu').filter(kategori='luar_biasa', is_active=True)
    
    context = {
        'dosen_tetap': dosen_tetap,
        'dosen_dtpr': dosen_dtpr,
        'dosen_lb': dosen_lb,
    }
    return render(request, 'akademik.html', context)


def kurikulum(request):
    """Halaman Kurikulum"""
    # 1. OBE K22 / K23
    try:
        kurikulum_obe = DocumentCategory.objects.get(slug='kurikulum-berbasis-obe-k22')
        docs_obe = kurikulum_obe.documents.filter(is_active=True)
    except DocumentCategory.DoesNotExist:
        docs_obe = []
    
    # 2. KKNI K17
    try:
        kurikulum_kkni = DocumentCategory.objects.get(slug='kurikulum-berbasis-kkni-k17')
        docs_kkni = kurikulum_kkni.documents.filter(is_active=True)
    except DocumentCategory.DoesNotExist:
        docs_kkni = []
    
    # 3. Mata Kuliah (List Mata Kuliah dan Semester)
    # Fetch all active dosen for name-to-nidn mapping
    all_dosen = Dosen.objects.filter(is_active=True)
    dosen_map = {d.nama.strip(): d.nidn for d in all_dosen}
    
    mata_kuliah_by_semester = {}
    # Fetch semesters 1-8 plus 0 (Pilihan)
    semester_list = [1, 2, 3, 4, 5, 6, 7, 8, 0]
    for i in semester_list:
        # Fetch courses for each semester
        courses = list(MataKuliah.objects.filter(semester=i).order_by('order', 'kode'))
        if i == 0 and not courses:
            continue
            
        # Process dosen_pengampu for each course
        for mk in courses:
            # New Many-to-Many logic
            mk.dosen_list = []
            for d in mk.dosen_pengampu.all():
                mk.dosen_list.append({
                    'nama': d.nama,
                    'nidn': d.nidn
                })
            
            # Fallback to legacy field if M2M is empty
            if not mk.dosen_list and mk.dosen_pengampu_legacy:
                import re
                separator_pattern = r'[,\n;]'
                raw_names = re.split(separator_pattern, mk.dosen_pengampu_legacy)
                
                clean_names = []
                for name in raw_names:
                    name = re.sub(r'^\s*\d+[\.\)]\s*', '', name.strip())
                    if name:
                        clean_names.append(name)
                
                names = sorted(list(set(clean_names)))
                for name in names:
                    mk.dosen_list.append({
                        'nama': name,
                        'nidn': dosen_map.get(name)
                    })
                
        mata_kuliah_by_semester[i] = courses
    
    # 4. RPS (Rencana Pembelajaran Semester)
    try:
        rps_cat = DocumentCategory.objects.get(slug='rencana-pembelajaran-semester-rps')
        docs_rps = rps_cat.documents.filter(is_active=True)
    except DocumentCategory.DoesNotExist:
        docs_rps = []
    
    # 5. Jadwal Perkuliahan (sinkronisasi dengan Master Mata Kuliah)
    # Mapping Master Mata Kuliah untuk sinkronisasi kode, nama, sks, dan semester
    mk_master_map = {mk.kode.strip().upper(): mk for mk in MataKuliah.objects.filter(is_active=True)}
    
    jadwal_items = JadwalPerkuliahanItem.objects.filter(is_active=True).order_by('semester', 'dosen', 'kode_mk')
    
    # {semester_key: {dosen_key: [items]}}
    jadwal_by_semester = {}
    
    # Semester labels mapping (from MataKuliah.SEMESTER_CHOICES)
    semester_labels = dict(MataKuliah.SEMESTER_CHOICES)
    
    for item in jadwal_items:
        # Sinkronisasi dengan Master Mata Kuliah jika ada kecocokan kode
        clean_kode = item.kode_mk.strip().upper()
        mk_master = mk_master_map.get(clean_kode)
        
        # Tentukan semester ID
        if mk_master:
            sem_id = mk_master.semester
            # Gunakan data master agar sinkron sempurna
            item.kode_mk = mk_master.kode
            item.nama_mk = mk_master.nama
            item.sks = mk_master.sks
        else:
            # Fallback jika tidak ada di master (mencoba parsing semester dari string)
            import re
            match = re.search(r'\d+', str(item.semester))
            sem_id = int(match.group()) if match else 99
        
        sem_label = semester_labels.get(sem_id, f"Semester {item.semester}")
        if sem_id == 0: sem_label = "Mata Kuliah Pilihan"
        
        sem_key = (sem_id, sem_label)
        if sem_key not in jadwal_by_semester:
            jadwal_by_semester[sem_key] = {}
            
        d_name = item.dosen or "Belum Ditentukan"
        d_nidn = item.kode_dosen or ""
        d_key = (d_name, d_nidn)
        
        if d_key not in jadwal_by_semester[sem_key]:
            jadwal_by_semester[sem_key][d_key] = []
            
        jadwal_by_semester[sem_key][d_key].append(item)

    # Urutkan semester (1, 2, 3... lalu 0, lalu 99)
    sorted_sem_keys = sorted(jadwal_by_semester.keys(), key=lambda x: (x[0] if x[0] > 0 else 90) if x[0] != 99 else 100)
    
    # Rebuild a sorted list of dictionaries [{sem_label, dosen_items: {dosen_key: [items]}}]
    # Or just keep it as a sorted dictionary for the template to iterate
    ordered_jadwal = {}
    for skey in sorted_sem_keys:
        ordered_jadwal[skey[1]] = jadwal_by_semester[skey]

    context = {
        'docs_obe': docs_obe,
        'docs_kkni': docs_kkni,
        'mata_kuliah_by_semester': mata_kuliah_by_semester,
        'docs_rps': docs_rps,
        'jadwal_by_semester': ordered_jadwal,
    }
    return render(request, 'kurikulum.html', context)


@staff_member_required
def import_jadwal_excel(request):
    """Import jadwal perkuliahan dari file Excel"""
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        batch_label = request.POST.get('batch_label', '').strip()
        clear_old = request.POST.get('clear_old')

        if not excel_file:
            messages.error(request, 'Silakan pilih file Excel.')
            return redirect('import_jadwal_excel')

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Format file tidak didukung. Gunakan file .xlsx atau .xls')
            return redirect('import_jadwal_excel')

        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True)
            ws = wb.active

            # Read header row and detect columns (case-insensitive keyword matching)
            headers = []
            for cell in ws[1]:
                val = str(cell.value or '').strip().upper()
                headers.append(val)

            # Smart column detection
            col_map = {}
            for idx, h in enumerate(headers):
                h_lower = h.lower()
                if 'smt' in h_lower or 'semester' in h_lower:
                    col_map['semester'] = idx
                elif 'kode' in h_lower:
                    col_map['kode'] = idx
                elif 'nama' in h_lower and 'mata' in h_lower:
                    col_map['nama'] = idx
                elif 'nama' in h_lower and 'kode' not in h_lower:
                    col_map['nama'] = idx
                elif 'sks' in h_lower:
                    col_map['sks'] = idx
                elif 'jadwal' in h_lower:
                    col_map['jadwal'] = idx
                elif 'dosen' in h_lower or 'pengampu' in h_lower:
                    col_map['dosen'] = idx

            required = ['semester', 'kode', 'jadwal']
            missing = [k for k in required if k not in col_map]
            if missing:
                messages.error(request, f'Kolom berikut tidak ditemukan di file Excel: {", ".join(missing)}. Header yang terdeteksi: {", ".join(headers)}')
                return redirect('import_jadwal_excel')

            # Clear old data if requested
            if clear_old:
                deleted_count = JadwalPerkuliahanItem.objects.all().delete()[0]
                messages.info(request, f'{deleted_count} data jadwal lama telah dihapus.')

            # Parse rows
            items_to_create = []
            row_count = 0
            error_rows = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row is None or all(c is None for c in row):
                    continue
                try:
                    smt_val = str(row[col_map['semester']] or '').strip()
                    if not smt_val:
                        continue
                    kode = str(row[col_map['kode']] or '').strip()
                    nama = str(row[col_map.get('nama', -1)] or '').strip() if 'nama' in col_map else ''
                    sks_val = row[col_map.get('sks', -1)] if 'sks' in col_map else 0
                    sks = int(sks_val) if sks_val else 0
                    jadwal = str(row[col_map['jadwal']] or '').strip()
                    ruangan = str(row[col_map.get('ruangan', -1)] or '').strip() if 'ruangan' in col_map else ''
                    dosen_str = str(row[col_map.get('dosen', -1)] or '').strip() if 'dosen' in col_map and col_map['dosen'] < len(row) else ''

                    if not kode:
                        continue

                    # Lookup for automatic data filling if missing in Excel
                    if not nama or sks == 0:
                        mk_match = MataKuliah.objects.filter(kode__iexact=kode).first()
                        if mk_match:
                            if not nama: nama = mk_match.nama
                            if sks == 0: sks = mk_match.sks

                    # Lookup Dosen for NIDN
                    kode_d_val = ""
                    if dosen_str:
                        d_match = Dosen.objects.filter(nama__icontains=dosen_str).first()
                        if d_match:
                            kode_d_val = d_match.nidn

                    items_to_create.append(JadwalPerkuliahanItem(
                        semester=smt_val,
                        kode_mk=kode,
                        nama_mk=nama,
                        sks=sks,
                        jadwal=jadwal,
                        ruangan=ruangan,
                        dosen=dosen_str,
                        kode_dosen=kode_d_val,
                        batch_label=batch_label,
                    ))
                    row_count += 1
                except Exception as e:
                    error_rows.append(f'Baris {row_idx}: {str(e)}')

            if items_to_create:
                JadwalPerkuliahanItem.objects.bulk_create(items_to_create)
                messages.success(request, f'✅ Berhasil mengimpor {row_count} data jadwal perkuliahan.')
            else:
                messages.warning(request, 'Tidak ada data valid yang ditemukan dalam file Excel.')

            if error_rows:
                messages.warning(request, f'⚠️ {len(error_rows)} baris bermasalah: {"; ".join(error_rows[:5])}')

            wb.close()
            return redirect('import_jadwal_excel')

        except Exception as e:
            messages.error(request, f'Gagal memproses file Excel: {str(e)}')
            return redirect('import_jadwal_excel')

    return render(request, 'admin/import_jadwal_excel.html')


def akreditasi(request):
    """Halaman Akreditasi"""
    akreditasi_list = Akreditasi.objects.filter(is_active=True)
    
    try:
        led = DocumentCategory.objects.get(slug='laporan-evaluasi-diri')
        led_docs = led.documents.filter(is_active=True)
    except DocumentCategory.DoesNotExist:
        led_docs = []
    
    try:
        lkps = DocumentCategory.objects.get(slug='lkps')
        lkps_docs = lkps.documents.filter(is_active=True)
    except DocumentCategory.DoesNotExist:
        lkps_docs = []
    
    context = {
        'akreditasi_list': akreditasi_list,
        'led_docs': led_docs,
        'lkps_docs': lkps_docs,
    }
    return render(request, 'akreditasi.html', context)


def kemahasiswaan(request):
    """Halaman Kemahasiswaan & Alumni"""
    prestasi = Prestasi.objects.all()[:10]
    alumni = Alumni.objects.all()[:12]
    
    try:
        buku_panduan = DocumentCategory.objects.get(slug='buku-panduan')
        panduan_docs = buku_panduan.documents.filter(is_active=True)
    except DocumentCategory.DoesNotExist:
        panduan_docs = []
    
    context = {
        'prestasi': prestasi,
        'alumni': alumni,
        'panduan_docs': panduan_docs,
    }
    return render(request, 'kemahasiswaan.html', context)


def layanan(request):
    """Halaman Layanan"""
    links = Link.objects.filter(is_active=True)
    penelitian = Penelitian.objects.all()[:10]
    job_careers = JobCareer.objects.filter(is_active=True)[:10]
    
    # Dokumen Pembimbing
    try:
        pa_cat = DocumentCategory.objects.get(slug='pembimbing-akademik')
        pa_docs = pa_cat.documents.filter(is_active=True)
    except DocumentCategory.DoesNotExist:
        pa_docs = []
    
    try:
        ps_cat = DocumentCategory.objects.get(slug='pembimbing-seminar')
        ps_docs = ps_cat.documents.filter(is_active=True)
    except DocumentCategory.DoesNotExist:
        ps_docs = []
    
    try:
        psk_cat = DocumentCategory.objects.get(slug='pembimbing-skripsi')
        psk_docs = psk_cat.documents.filter(is_active=True)
    except DocumentCategory.DoesNotExist:
        psk_docs = []
    
    context = {
        'links': links,
        'penelitian': penelitian,
        'job_careers': job_careers,
        'pa_docs': pa_docs,
        'ps_docs': ps_docs,
        'psk_docs': psk_docs,
    }
    return render(request, 'layanan.html', context)


def media_informasi(request):
    """Halaman Media & Informasi"""
    news_list = News.objects.filter(is_published=True)
    gallery_photos = Gallery.objects.filter(is_active=True, media_type='image')
    gallery_videos = Gallery.objects.filter(is_active=True, media_type='video')
    
    context = {
        'news_list': news_list,
        'gallery_photos': gallery_photos,
        'gallery_videos': gallery_videos,
    }
    return render(request, 'media_informasi.html', context)


def news_detail(request, slug):
    """Detail Berita"""
    news = get_object_or_404(News, slug=slug, is_published=True)
    news.increment_views()
    recent_news = News.objects.filter(is_published=True).exclude(id=news.id)[:5]
    
    context = {
        'news': news,
        'recent_news': recent_news,
    }
    return render(request, 'news_detail.html', context)


def documents_by_category(request, slug):
    """Daftar dokumen per kategori"""
    category = get_object_or_404(DocumentCategory, slug=slug)
    documents = category.documents.filter(is_active=True)
    
    context = {
        'category': category,
        'documents': documents,
    }
    return render(request, 'documents.html', context)


def download_document(request, doc_id):
    """Download atau tampilkan dokumen"""
    document = get_object_or_404(Document, id=doc_id, is_active=True)
    view_type = request.GET.get('view', 'download')
    
    if os.path.exists(document.file.path):
        document.increment_download()
        response = FileResponse(open(document.file.path, 'rb'))
        
        disposition = 'inline' if view_type == 'inline' else 'attachment'
        filename = os.path.basename(document.file.name)
        response['Content-Disposition'] = f'{disposition}; filename="{filename}"'
        
        # Ensure content type is set for PDFs to open in browser
        if filename.lower().endswith('.pdf'):
            response['Content-Type'] = 'application/pdf'
            
        return response
    else:
        raise Http404("File tidak ditemukan")


def search(request):
    """Pencarian"""
    query = request.GET.get('q', '')
    results = []
    
    if query:
        # Search in News
        news_results = News.objects.filter(
            Q(title__icontains=query) | Q(content__icontains=query),
            is_published=True
        )
        
        # Search in Documents
        doc_results = Document.objects.filter(
            Q(title__icontains=query) | Q(description__icontains=query),
            is_active=True
        )
        
        # Search in Dosen
        dosen_results = Dosen.objects.filter(
            Q(nama__icontains=query) | Q(bidang_keahlian__icontains=query),
            is_active=True
        )
        
        results = {
            'news': news_results,
            'documents': doc_results,
            'dosen': dosen_results,
            'query': query
        }
    
    return render(request, 'search.html', results)


# Admin Views
@login_required
def admin_dashboard(request):
    """Dashboard Admin"""
    stats = {
        'total_news': News.objects.count(),
        'total_documents': Document.objects.count(),
        'total_dosen': Dosen.objects.count(),
        'total_alumni': Alumni.objects.count(),
        'total_survei_kurikulum': SurveyKurikulum.objects.count(),
    }
    
    recent_news = News.objects.all()[:5]
    recent_docs = Document.objects.all()[:5]
    
    context = {
        'stats': stats,
        'recent_news': recent_news,
        'recent_docs': recent_docs,
    }
    return render(request, 'admin/dashboard.html', context)


def user_login(request):
    """Login"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, 'Login berhasil!')
            return redirect('admin_dashboard')
        else:
            messages.error(request, 'Username atau password salah!')
    
    return render(request, 'login.html')


def user_logout(request):
    """Logout"""
    logout(request)
    messages.success(request, 'Logout berhasil!')
    return redirect('home')

def survey_view(request):
    """Halaman Survei VMTS"""
    if request.method == 'POST':
        # Identitas
        status = request.POST.get('status')
        lama_bergabung = request.POST.get('lama')
        
        # Sosialisasi
        pengetahuan_visi = request.POST.get('pengetahuan_visi')
        sumber_informasi = request.POST.getlist('sumber')  # Multiple choice
        frekuensi_sosialisasi = request.POST.get('frekuensi')
        
        # Pemahaman & Analisis
        tingkat_paham = request.POST.get('tingkat_paham')
        aspek_tercermin = request.POST.get('aspek_tercermin')
        bidang_tercermin = request.POST.getlist('bidang')  # Multiple choice
        dukungan_atmosfer = request.POST.get('dukungan_atmosfer')
        perlu_perbaikan = request.POST.get('perlu_perbaikan')
        saran_kritik = request.POST.get('saran')
        
        # Save to database
        response = SurveyResponse.objects.create(
            status=status,
            lama_bergabung=lama_bergabung,
            pengetahuan_visi=pengetahuan_visi,
            sumber_informasi=sumber_informasi,
            frekuensi_sosialisasi=frekuensi_sosialisasi,
            tingkat_paham=tingkat_paham,
            aspek_tercermin=aspek_tercermin,
            bidang_tercermin=bidang_tercermin,
            dukungan_atmosfer=dukungan_atmosfer,
            perlu_perbaikan=perlu_perbaikan,
            saran_kritik=saran_kritik
        )
        
        messages.success(request, 'Terima kasih atas partisipasi Anda dalam survei VMTS!')
        return redirect('layanan')
        
    return render(request, 'survey.html')

def submit_survey_lulusan(request):
    """Halaman Survei Kepuasan Pengguna Lulusan"""
    aspects = [
        {'name': 'integritas', 'label': 'Integritas (Etika dan Moral)'},
        {'name': 'keahlian_bidang', 'label': 'Keahlian pada Bidang Ilmu (Kompetensi Utama)'},
        {'name': 'bahasa_inggris', 'label': 'Kemampuan Bahasa Inggris'},
        {'name': 'penggunaan_it', 'label': 'Penggunaan Teknologi Informasi (IT)'},
        {'name': 'komunikasi', 'label': 'Kemampuan Komunikasi'},
        {'name': 'kerjasama_tim', 'label': 'Kerjasama Tim'},
        {'name': 'pengembangan_diri', 'label': 'Pengembangan Diri'},
    ]
    
    if request.method == 'POST':
        SurveyKepuasanLulusan.objects.create(
            email=request.POST.get('email'),
            nama_responden=request.POST.get('nama_responden'),
            nama_instansi=request.POST.get('nama_instansi'),
            jabatan=request.POST.get('jabatan'),
            integritas=request.POST.get('integritas'),
            keahlian_bidang=request.POST.get('keahlian_bidang'),
            bahasa_inggris=request.POST.get('bahasa_inggris'),
            penggunaan_it=request.POST.get('penggunaan_it'),
            komunikasi=request.POST.get('komunikasi'),
            kerjasama_tim=request.POST.get('kerjasama_tim'),
            pengembangan_diri=request.POST.get('pengembangan_diri'),
            waktu_tunggu=request.POST.get('waktu_tunggu')
        )
        messages.success(request, 'Terima kasih! Survei kepuasan pelanggan telah berhasil dikirim.')
        return redirect('layanan')
        
    return render(request, 'survey/form_pengguna.html', {'aspects': aspects})

@login_required
def export_survey_excel(request):
    """Export data survei ke Excel"""
    surveys = SurveyKepuasanLulusan.objects.all().values()
    df = pd.DataFrame(list(surveys))
    
    # Excel does not support timezone-aware datetimes
    if 'created_at' in df.columns:
        df['created_at'] = df['created_at'].dt.tz_localize(None)
    
    # Format response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="survey_kepuasan_lulusan.xlsx"'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Survey Data')
        
    return response

@login_required
def survey_stats_view(request):
    """Dasbor Statistik Survei (Infografis)"""
    # Mapping Skala Likert ke Angka untuk Rata-rata
    skala_map = {
        'Sangat Baik': 4,
        'Baik': 3,
        'Cukup': 2,
        'Kurang': 1
    }
    
    surveys = SurveyKepuasanLulusan.objects.all()
    total_responden = surveys.count()
    
    if total_responden == 0:
        return render(request, 'admin/survey_stats.html', {'empty': True})
        
    # Kalkulasi Rata-rata tiap aspek
    aspek_config = {
        'integritas': 'Integritas',
        'keahlian_bidang': 'Keahlian Bidang',
        'bahasa_inggris': 'Bahasa Inggris',
        'penggunaan_it': 'Penggunaan IT',
        'komunikasi': 'Komunikasi',
        'kerjasama_tim': 'Kerjasama Tim',
        'pengembangan_diri': 'Pengembangan Diri'
    }
    
    stats_data = [] # List of dicts for easier template rendering
    chart_labels = []
    chart_values = []
    
    for field, label in aspek_config.items():
        # Get count for each scale
        counts = surveys.values(field).annotate(total=Count(field))
        total_score = 0
        for item in counts:
            score = skala_map.get(item[field], 0)
            total_score += score * item['total']
        
        avg_score = round(total_score / total_responden, 2) if total_responden > 0 else 0
        stats_data.append({
            'label': label,
            'value': avg_score
        })
        chart_labels.append(label)
        chart_values.append(avg_score)

    # Waktu Tunggu Pie Chart Data
    wt_data = surveys.values('waktu_tunggu').annotate(total=Count('waktu_tunggu'))
    
    context = {
        'stats_data': stats_data,
        'chart_labels': chart_labels,
        'chart_values': chart_values,
        'wt_data': wt_data,
        'total_responden': total_responden,
        'empty': False
    }
    return render(request, 'admin/survey_stats.html', context)

@login_required
def export_vmts_excel(request):
    """Export data survei VMTS ke Excel"""
    responses = SurveyResponse.objects.all().values()
    df = pd.DataFrame(list(responses))
    
    # Excel does not support timezone-aware datetimes
    if 'created_at' in df.columns:
        df['created_at'] = df['created_at'].dt.tz_localize(None)
    
    # Format response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="survey_vmts_responses.xlsx"'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='VMTS Data')
        
    return response

@login_required
def vmts_stats_view(request):
    """Dasbor Statistik Survei VMTS (Infografis)"""
    responses = SurveyResponse.objects.all()
    total_responden = responses.count()
    
    if total_responden == 0:
        return render(request, 'admin/vmts_stats.html', {'empty': True})
        
    # Agregasi data untuk grafik
    status_dist = responses.values('status').annotate(total=Count('status'))
    tingkat_paham_dist = responses.values('tingkat_paham').annotate(total=Count('tingkat_paham'))
    frekuensi_sos_dist = responses.values('frekuensi_sosialisasi').annotate(total=Count('frekuensi_sosialisasi'))
    
    context = {
        'total_responden': total_responden,
        'status_dist': status_dist,
        'tingkat_paham_dist': tingkat_paham_dist,
        'frekuensi_sos_dist': frekuensi_sos_dist,
        'empty': False
    }
    return render(request, 'admin/vmts_stats.html', context)

@login_required
def kurikulum_stats_view(request):
    """Dasbor Statistik Survei Kurikulum (Infografis)"""
    surveys = SurveyKurikulum.objects.all()
    total_responden = surveys.count()
    
    if total_responden == 0:
        return render(request, 'admin/kurikulum_stats.html', {'empty': True})
        
    # Group by category
    category_counts = surveys.values('kategori_responden').annotate(total=Count('kategori_responden'))
    
    # 1. Alumni CPL Stats
    alumni_surveys = surveys.filter(kategori_responden='alumni')
    alumni_count = alumni_surveys.count()
    alumni_stats = []
    
    if alumni_count > 0:
        cpl_fields = {
            'alumni_cpl_sikap': 'Sikap & Etika',
            'alumni_cpl_konsep': 'Konsep SI',
            'alumni_cpl_teknis': 'Teknis Pemrograman',
            'alumni_cpl_data': 'Manajemen Data',
            'alumni_cpl_tata_kelola': 'Tata Kelola & Audit',
            'alumni_cpl_techno': 'Technopreneurship'
        }
        
        for field, label in cpl_fields.items():
            total_score = 0
            count_valid = 0
            for s in alumni_surveys:
                val = s.responses_data.get(field)
                if val:
                    try:
                        total_score += int(val)
                        count_valid += 1
                    except ValueError:
                        pass
            
            avg = round(total_score / count_valid, 2) if count_valid > 0 else 0
            alumni_stats.append({'label': label, 'value': avg})

    # 2. Pengguna Lulusan Stats
    user_surveys = surveys.filter(kategori_responden='pengguna')
    user_count = user_surveys.count()
    user_stats = []
    
    if user_count > 0:
        eval_fields = {
            'user_eval_integritas': 'Integritas',
            'user_eval_keahlian_si': 'Keahlian SI',
            'user_eval_keahlian_teknis': 'Keahlian Teknis',
            'user_eval_soft_skill': 'Soft Skill',
            'user_eval_adaptasi': 'Adaptasi'
        }
        
        for field, label in eval_fields.items():
            total_score = 0
            count_valid = 0
            for s in user_surveys:
                val = s.responses_data.get(field)
                if val:
                    try:
                        total_score += int(val)
                        count_valid += 1
                    except ValueError:
                        pass
            
            avg = round(total_score / count_valid, 2) if count_valid > 0 else 0
            user_stats.append({'label': label, 'value': avg})

    context = {
        'total_responden': total_responden,
        'category_counts': category_counts,
        'alumni_stats': alumni_stats,
        'user_stats': user_stats,
        'empty': False
    }
    return render(request, 'admin/kurikulum_stats.html', context)

def submit_survey_kurikulum(request):
    """Halaman Survei Evaluasi Kurikulum"""
    if request.method == 'POST':
        # List of metadata fields to exclude from instrument data
        meta_fields = ['csrfmiddlewaretoken', 'nama_responden', 'kategori_responden', 'instansi', 'jabatan', 'email']
        
        # Collect all other POST data into a dict
        responses = {k: v for k, v in request.POST.items() if k not in meta_fields and v}
        
        SurveyKurikulum.objects.create(
            nama_responden=request.POST.get('nama_responden'),
            kategori_responden=request.POST.get('kategori_responden'),
            instansi=request.POST.get('instansi'),
            jabatan=request.POST.get('jabatan') or '',
            email=request.POST.get('email') or '',
            responses_data=responses
        )
        messages.success(request, 'Terima kasih atas partisipasi Anda dalam survei kurikulum!')
        return redirect('layanan')
        
    return render(request, 'survey/form_kurikulum.html')
